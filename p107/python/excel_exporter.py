import os
from typing import List, Dict, Optional
from dataclasses import dataclass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@dataclass
class DBCSignalForExport:
    name: str
    start_bit: int
    bit_length: int
    is_signed: bool
    is_big_endian: bool
    scale: float
    offset: float
    min_value: Optional[float]
    max_value: Optional[float]
    unit: str
    receiver: str = ""
    comment: str = ""


@dataclass
class DBCMessageForExport:
    can_id: int
    name: str
    dlc: int
    sender: str = ""
    comment: str = ""
    signals: List[DBCSignalForExport] = None
    cycle_time: int = 0

    def __post_init__(self):
        if self.signals is None:
            self.signals = []


class ExcelExporter:
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
        
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        self.subheader_font = Font(bold=True, size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')

    def export_dbc_to_excel(self, messages: List[DBCMessageForExport], output_path: str,
                            project_name: str = "CAN Analysis") -> str:
        wb = Workbook()
        wb.remove(wb.active)

        self._create_overview_sheet(wb, messages, project_name)
        self._create_messages_sheet(wb, messages)
        self._create_signals_sheet(wb, messages)
        self._create_signal_details_sheet(wb, messages)

        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        
        return output_path

    def _create_overview_sheet(self, wb: Workbook, messages: List[DBCMessageForExport], 
                               project_name: str):
        ws = wb.create_sheet("Overview")

        ws['A1'] = "CAN Database Overview"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')

        ws['A3'] = "Project Name:"
        ws['B3'] = project_name
        ws['A4'] = "Export Date:"
        ws['B4'] = "=NOW()"
        ws['A5'] = "Total Messages:"
        ws['B5'] = len(messages)
        
        total_signals = sum(len(msg.signals) for msg in messages)
        ws['A6'] = "Total Signals:"
        ws['B6'] = total_signals

        headers = ['CAN ID (Hex)', 'Message Name', 'DLC', 'Signal Count', 'Cycle Time (ms)', 'Sender']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.border

        for row, msg in enumerate(messages, 10):
            ws.cell(row=row, column=1, value=f"0x{msg.can_id:03X}").border = self.border
            ws.cell(row=row, column=2, value=msg.name).border = self.border
            ws.cell(row=row, column=3, value=msg.dlc).border = self.border
            ws.cell(row=row, column=4, value=len(msg.signals)).border = self.border
            ws.cell(row=row, column=5, value=msg.cycle_time).border = self.border
            ws.cell(row=row, column=6, value=msg.sender).border = self.border

        self._auto_fit_columns(ws, 6)

    def _create_messages_sheet(self, wb: Workbook, messages: List[DBCMessageForExport]):
        ws = wb.create_sheet("Messages")

        headers = ['CAN ID (Hex)', 'CAN ID (Dec)', 'Message Name', 'DLC', 'Sender', 
                   'Cycle Time (ms)', 'Comment', 'Signal Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.border

        for row, msg in enumerate(messages, 2):
            ws.cell(row=row, column=1, value=f"0x{msg.can_id:03X}").border = self.border
            ws.cell(row=row, column=2, value=msg.can_id).border = self.border
            ws.cell(row=row, column=3, value=msg.name).border = self.border
            ws.cell(row=row, column=4, value=msg.dlc).border = self.border
            ws.cell(row=row, column=5, value=msg.sender).border = self.border
            ws.cell(row=row, column=6, value=msg.cycle_time).border = self.border
            ws.cell(row=row, column=7, value=msg.comment).border = self.border
            ws.cell(row=row, column=8, value=len(msg.signals)).border = self.border

        self._auto_fit_columns(ws, 8)

    def _create_signals_sheet(self, wb: Workbook, messages: List[DBCMessageForExport]):
        ws = wb.create_sheet("Signals")

        headers = ['Message', 'CAN ID', 'Signal Name', 'Start Bit', 'Bit Length', 
                   'Byte Order', 'Value Type', 'Scale', 'Offset', 'Min', 'Max', 'Unit', 'Receiver']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.border

        row = 2
        for msg in messages:
            for signal in msg.signals:
                ws.cell(row=row, column=1, value=msg.name).border = self.border
                ws.cell(row=row, column=2, value=f"0x{msg.can_id:03X}").border = self.border
                ws.cell(row=row, column=3, value=signal.name).border = self.border
                ws.cell(row=row, column=4, value=signal.start_bit).border = self.border
                ws.cell(row=row, column=5, value=signal.bit_length).border = self.border
                ws.cell(row=row, column=6, value="Big Endian" if signal.is_big_endian else "Little Endian").border = self.border
                ws.cell(row=row, column=7, value="Signed" if signal.is_signed else "Unsigned").border = self.border
                ws.cell(row=row, column=8, value=signal.scale).border = self.border
                ws.cell(row=row, column=9, value=signal.offset).border = self.border
                ws.cell(row=row, column=10, value=signal.min_value if signal.min_value is not None else "").border = self.border
                ws.cell(row=row, column=11, value=signal.max_value if signal.max_value is not None else "").border = self.border
                ws.cell(row=row, column=12, value=signal.unit).border = self.border
                ws.cell(row=row, column=13, value=signal.receiver).border = self.border
                row += 1

        self._auto_fit_columns(ws, 13)

    def _create_signal_details_sheet(self, wb: Workbook, messages: List[DBCMessageForExport]):
        ws = wb.create_sheet("Signal Details")

        for msg_idx, msg in enumerate(messages):
            start_row = 1 + msg_idx * 20

            title_cell = ws.cell(row=start_row, column=1, 
                                value=f"{msg.name} (0x{msg.can_id:03X})")
            title_cell.font = Font(bold=True, size=14, color="4472C4")
            ws.merge_cells(start_row=start_row, start_column=1, 
                          end_row=start_row, end_column=10)

            ws.cell(row=start_row + 2, column=1, value="DLC:").font = self.subheader_font
            ws.cell(row=start_row + 2, column=2, value=msg.dlc)
            ws.cell(row=start_row + 2, column=4, value="Sender:").font = self.subheader_font
            ws.cell(row=start_row + 2, column=5, value=msg.sender)
            ws.cell(row=start_row + 3, column=1, value="Cycle Time:").font = self.subheader_font
            ws.cell(row=start_row + 3, column=2, value=f"{msg.cycle_time} ms")

            if msg.signals:
                signal_headers = ['Signal Name', 'Start Bit', 'Length', 'Byte Order', 'Type',
                                  'Scale', 'Offset', 'Min', 'Max', 'Unit']
                for col, header in enumerate(signal_headers, 1):
                    cell = ws.cell(row=start_row + 5, column=col, value=header)
                    cell.fill = self.subheader_fill
                    cell.font = self.subheader_font
                    cell.alignment = self.center_align
                    cell.border = self.border

                for sig_row, signal in enumerate(msg.signals, start_row + 6):
                    ws.cell(row=sig_row, column=1, value=signal.name).border = self.border
                    ws.cell(row=sig_row, column=2, value=signal.start_bit).border = self.border
                    ws.cell(row=sig_row, column=3, value=signal.bit_length).border = self.border
                    ws.cell(row=sig_row, column=4, value="BE" if signal.is_big_endian else "LE").border = self.border
                    ws.cell(row=sig_row, column=5, value="S" if signal.is_signed else "U").border = self.border
                    ws.cell(row=sig_row, column=6, value=signal.scale).border = self.border
                    ws.cell(row=sig_row, column=7, value=signal.offset).border = self.border
                    ws.cell(row=sig_row, column=8, value=signal.min_value if signal.min_value is not None else "").border = self.border
                    ws.cell(row=sig_row, column=9, value=signal.max_value if signal.max_value is not None else "").border = self.border
                    ws.cell(row=sig_row, column=10, value=signal.unit).border = self.border

        self._auto_fit_columns(ws, 10)

    def _auto_fit_columns(self, ws, num_cols: int):
        for col in range(1, num_cols + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    if cell.value:
                        try:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                        except:
                            pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


def export_dbc_from_database(project_id: int, output_path: str, db) -> str:
    from database import Database
    
    db_instance = db if db else Database()
    signals_by_can_id = db_instance.get_signals(project_id)
    can_ids = db_instance.get_unique_can_ids(project_id)
    project = db_instance.get_project(project_id)
    
    messages = []
    for can_id in can_ids:
        dbc_msg = DBCMessageForExport(
            can_id=can_id,
            name=f"Message_0x{can_id:03X}",
            dlc=8
        )
        
        signals = signals_by_can_id.get(can_id, [])
        for sig in signals:
            dbc_signal = DBCSignalForExport(
                name=sig['name'],
                start_bit=sig['start_bit'],
                bit_length=sig['bit_length'],
                is_signed=sig['is_signed'],
                is_big_endian=sig['is_big_endian'],
                scale=sig['scale'],
                offset=sig['offset'],
                min_value=None,
                max_value=None,
                unit=sig.get('unit', '')
            )
            dbc_msg.signals.append(dbc_signal)
        
        messages.append(dbc_msg)
    
    exporter = ExcelExporter()
    project_name = project['name'] if project else "CAN Analysis"
    return exporter.export_dbc_to_excel(messages, output_path, project_name)


if __name__ == '__main__':
    if OPENPYXL_AVAILABLE:
        test_messages = [
            DBCMessageForExport(
                can_id=0x100,
                name="EngineData",
                dlc=8,
                sender="ECU_Engine",
                cycle_time=10,
                signals=[
                    DBCSignalForExport(
                        name="EngineSpeed",
                        start_bit=0,
                        bit_length=16,
                        is_signed=False,
                        is_big_endian=False,
                        scale=1.0,
                        offset=0.0,
                        min_value=0.0,
                        max_value=8000.0,
                        unit="rpm"
                    ),
                    DBCSignalForExport(
                        name="EngineTemp",
                        start_bit=16,
                        bit_length=8,
                        is_signed=True,
                        is_big_endian=False,
                        scale=1.0,
                        offset=-40.0,
                        min_value=-40.0,
                        max_value=215.0,
                        unit="C"
                    )
                ]
            ),
            DBCMessageForExport(
                can_id=0x200,
                name="VehicleData",
                dlc=8,
                sender="BCM",
                cycle_time=50,
                signals=[
                    DBCSignalForExport(
                        name="VehicleSpeed",
                        start_bit=0,
                        bit_length=16,
                        is_signed=False,
                        is_big_endian=False,
                        scale=0.01,
                        offset=0.0,
                        min_value=0.0,
                        max_value=655.35,
                        unit="km/h"
                    )
                ]
            )
        ]

        exporter = ExcelExporter()
        output_path = "../output/test_dbc_export.xlsx"
        exporter.export_dbc_to_excel(test_messages, output_path, "Test Project")
        print(f"Excel exported to: {output_path}")
    else:
        print("openpyxl not available. Install with: pip install openpyxl")
