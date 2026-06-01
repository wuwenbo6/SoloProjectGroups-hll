from fastapi import FastAPI, File, UploadFile, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
import os
import uuid
import io
from datetime import datetime, timedelta
import json
import cv2

from .database import engine, get_db, Base
from . import models, schemas
from .detector import detector
from .individual_recognizer import recognizer_manager

Base.metadata.create_all(bind=engine)

app = FastAPI(title="野生动物检测系统")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")


@app.post("/api/photos/upload", response_model=schemas.PhotoWithDetections)
async def upload_photo(
    file: UploadFile = File(...),
    camera_id: Optional[str] = Query(None),
    location: Optional[str] = Query(None),
    enable_individual_recognition: bool = Query(True),
    db: Session = Depends(get_db)
):
    file_ext = os.path.splitext(file.filename)[1]
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, unique_filename)

    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)

    db_photo = models.Photo(
        filename=file.filename,
        file_path=f"/uploads/{unique_filename}",
        camera_id=camera_id,
        location=location
    )
    db.add(db_photo)
    db.commit()
    db.refresh(db_photo)

    detection_results = detector.detect(file_path)

    image = cv2.imread(file_path)
    species_counts = {}

    for result in detection_results:
        species = result["species"]
        count = result["count"]
        confidence = result["confidence"]
        
        if species not in species_counts:
            species_counts[species] = {
                "total_count": count,
                "processed_count": 0,
                "base_detection_added": False
            }

    if enable_individual_recognition and image is not None:
        try:
            from ultralytics import YOLO
            temp_model = YOLO("yolov8m.pt")
            results = temp_model(image, conf=0.25, verbose=False)
            
            for result in results:
                for box in result.boxes:
                    class_id = int(box.cls[0])
                    if class_id in detector.animal_classes:
                        species = detector.animal_classes[class_id]
                        if species in species_counts and species_counts[species]["processed_count"] < species_counts[species]["total_count"]:
                            x1, y1, x2, y2 = map(int, box.xyxy[0].cpu().numpy())
                            bbox = (x1, y1, x2, y2)
                            
                            detection, individual = recognizer_manager.process_detection(
                                db, species, image, bbox, db_photo.id
                            )
                            detection.confidence = f"{float(box.conf[0]):.2f}"
                            db.add(detection)
                            species_counts[species]["processed_count"] += 1
                            species_counts[species]["base_detection_added"] = True
        except Exception as e:
            print(f"个体识别处理出错: {e}")

    for species, data in species_counts.items():
        if not data["base_detection_added"]:
            for result in detection_results:
                if result["species"] == species:
                    db_detection = models.Detection(
                        photo_id=db_photo.id,
                        species=species,
                        count=data["total_count"],
                        confidence=result["confidence"]
                    )
                    db.add(db_detection)
                    break

    db.commit()
    db.refresh(db_photo)

    return db_photo


@app.get("/api/photos", response_model=List[schemas.PhotoWithDetections])
def get_photos(
    species: Optional[str] = Query(None),
    individual_id: Optional[str] = Query(None),
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    query = db.query(models.Photo).order_by(models.Photo.upload_time.desc())

    if species:
        query = query.join(models.Detection).filter(
            func.lower(models.Detection.species) == func.lower(species)
        )
    
    if individual_id:
        query = query.join(models.Detection).join(models.Individual).filter(
            models.Individual.individual_id == individual_id
        )

    photos = query.offset(skip).limit(limit).all()
    return photos


@app.get("/api/photos/{photo_id}", response_model=schemas.PhotoWithDetections)
def get_photo(photo_id: int, db: Session = Depends(get_db)):
    photo = db.query(models.Photo).filter(models.Photo.id == photo_id).first()
    if photo is None:
        raise HTTPException(status_code=404, detail="Photo not found")
    return photo


@app.get("/api/species")
def get_species_list(db: Session = Depends(get_db)):
    species_list = db.query(
        models.Detection.species,
        func.count(models.Detection.id).label("detection_count")
    ).group_by(models.Detection.species).all()

    return [
        {"species": item.species, "count": item.detection_count}
        for item in species_list
    ]


@app.delete("/api/photos/{photo_id}")
def delete_photo(photo_id: int, db: Session = Depends(get_db)):
    photo = db.query(models.Photo).filter(models.Photo.id == photo_id).first()
    if photo is None:
        raise HTTPException(status_code=404, detail="Photo not found")

    db.query(models.Detection).filter(models.Detection.photo_id == photo_id).delete()
    db.delete(photo)
    db.commit()

    return {"message": "Photo deleted successfully"}


@app.get("/api/individuals", response_model=List[schemas.Individual])
def get_individuals(
    species: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    query = db.query(models.Individual)
    if species:
        query = query.filter(models.Individual.species == species)
    
    individuals = query.order_by(models.Individual.sighting_count.desc()).all()
    return individuals


@app.get("/api/individuals/{individual_id}", response_model=schemas.IndividualWithDetections)
def get_individual(individual_id: str, db: Session = Depends(get_db)):
    individual = db.query(models.Individual).filter(
        models.Individual.individual_id == individual_id
    ).first()
    if individual is None:
        raise HTTPException(status_code=404, detail="Individual not found")
    return individual


@app.get("/api/recapture-rate", response_model=List[schemas.RecaptureRate])
def get_recapture_rate(
    species: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    return recognizer_manager.get_recapture_rate(db, species)


@app.get("/api/activity/heatmap", response_model=schemas.ActivityHeatmapData)
def get_activity_heatmap(
    species: str = Query(..., description="物种名称"),
    days: int = Query(7, ge=1, le=30, description="统计天数"),
    db: Session = Depends(get_db)
):
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)
    
    heatmap = [[0 for _ in range(24)] for _ in range(days)]
    hour_labels = [f"{h:02d}:00" for h in range(24)]
    day_labels = [(end_date - timedelta(days=i)).strftime("%m-%d") for i in range(days-1, -1, -1)]
    
    detections = db.query(models.Detection).join(models.Photo).filter(
        models.Detection.species == species,
        models.Photo.upload_time >= start_date
    ).all()
    
    for detection in detections:
        photo = detection.photo
        if photo and photo.upload_time:
            day_idx = (end_date.date() - photo.upload_time.date()).days
            if 0 <= day_idx < days:
                hour = photo.upload_time.hour
                heatmap[days - 1 - day_idx][hour] += 1
    
    return {
        "species": species,
        "heatmap": heatmap,
        "hour_labels": hour_labels,
        "day_labels": day_labels
    }


@app.get("/api/activity/species", response_model=List[schemas.SpeciesActivity])
def get_species_activity(
    species: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    query = db.query(models.Detection).join(models.Photo)
    if species:
        query = query.filter(models.Detection.species == species)
    
    detections = query.all()
    
    species_activity = {}
    for detection in detections:
        sp = detection.species
        if sp not in species_activity:
            species_activity[sp] = {str(h): 0 for h in range(24)}
        
        if detection.photo and detection.photo.upload_time:
            hour = detection.photo.upload_time.hour
            species_activity[sp][str(hour)] += 1
    
    result = []
    for sp, hourly_counts in species_activity.items():
        peak_hour = max(hourly_counts.items(), key=lambda x: x[1])[0]
        peak_count = hourly_counts[peak_hour]
        result.append({
            "species": sp,
            "hourly_counts": hourly_counts,
            "peak_hour": f"{int(peak_hour):02d}:00",
            "peak_count": peak_count
        })
    
    return result


@app.post("/api/report/export")
async def export_report(
    request: schemas.ExportReportRequest,
    db: Session = Depends(get_db)
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Please install it with: pip install openpyxl")

    wb = openpyxl.Workbook()
    
    ws_summary = wb.active
    ws_summary.title = "检测概览"
    
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    headers = ["统计项", "数值"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    total_photos = db.query(models.Photo).count()
    total_detections = db.query(models.Detection).count()
    species_list = db.query(models.Detection.species).distinct().count()
    
    summary_data = [
        ["总照片数", total_photos],
        ["总检测数", total_detections],
        ["物种数量", species_list],
        ["报告生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    ]
    
    for row, (key, value) in enumerate(summary_data, 2):
        ws_summary.cell(row=row, column=1, value=key)
        ws_summary.cell(row=row, column=2, value=value)
    
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 30

    if request.include_individuals:
        ws_individuals = wb.create_sheet("个体识别统计")
        
        headers = ["物种", "个体ID", "首次发现时间", "最后发现时间", "捕获次数"]
        for col, header in enumerate(headers, 1):
            cell = ws_individuals.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        individuals = db.query(models.Individual).all()
        for row, ind in enumerate(individuals, 2):
            ws_individuals.cell(row=row, column=1, value=ind.species)
            ws_individuals.cell(row=row, column=2, value=ind.individual_id)
            ws_individuals.cell(row=row, column=3, value=ind.first_seen.strftime("%Y-%m-%d %H:%M") if ind.first_seen else "")
            ws_individuals.cell(row=row, column=4, value=ind.last_seen.strftime("%Y-%m-%d %H:%M") if ind.last_seen else "")
            ws_individuals.cell(row=row, column=5, value=ind.sighting_count)
        
        for col in range(1, 6):
            ws_individuals.column_dimensions[chr(64 + col)].width = 20

    ws_photos = wb.create_sheet("检测详情")
    headers = ["照片文件名", "上传时间", "相机ID", "位置", "检测物种", "数量", "置信度", "个体ID"]
    for col, header in enumerate(headers, 1):
        cell = ws_photos.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    query = db.query(models.Photo).order_by(models.Photo.upload_time.desc())
    if request.start_date:
        query = query.filter(models.Photo.upload_time >= request.start_date)
    if request.end_date:
        query = query.filter(models.Photo.upload_time <= request.end_date)
    
    photos = query.all()
    row = 2
    for photo in photos:
        for detection in photo.detections:
            if request.species and detection.species != request.species:
                continue
            ws_photos.cell(row=row, column=1, value=photo.filename)
            ws_photos.cell(row=row, column=2, value=photo.upload_time.strftime("%Y-%m-%d %H:%M:%S") if photo.upload_time else "")
            ws_photos.cell(row=row, column=3, value=photo.camera_id or "")
            ws_photos.cell(row=row, column=4, value=photo.location or "")
            ws_photos.cell(row=row, column=5, value=detection.species)
            ws_photos.cell(row=row, column=6, value=detection.count)
            ws_photos.cell(row=row, column=7, value=detection.confidence)
            individual = db.query(models.Individual).filter(models.Individual.id == detection.individual_id).first()
            ws_photos.cell(row=row, column=8, value=individual.individual_id if individual else "")
            row += 1
    
    for col in range(1, 9):
        ws_photos.column_dimensions[chr(64 + col)].width = 20

    if request.include_activity:
        ws_activity = wb.create_sheet("活动节律")
        headers = ["物种", "时段", "检测数量"]
        for col, header in enumerate(headers, 1):
            cell = ws_activity.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        activity_data = get_species_activity(request.species, db)
        row = 2
        for sp_data in activity_data:
            for hour, count in sp_data["hourly_counts"].items():
                if count > 0:
                    ws_activity.cell(row=row, column=1, value=sp_data["species"])
                    ws_activity.cell(row=row, column=2, value=f"{int(hour):02d}:00")
                    ws_activity.cell(row=row, column=3, value=count)
                    row += 1
        
        ws_activity.column_dimensions['A'].width = 15
        ws_activity.column_dimensions['B'].width = 15
        ws_activity.column_dimensions['C'].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"wildlife_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
